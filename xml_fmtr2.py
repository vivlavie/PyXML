from xml.etree.ElementTree import parse

from openpyxl import load_workbook
import time
def remove_html_markup(s):
    tag = False
    quote = False
    out = ""
    if s != None:
        for c in s:
                if c == '<' and not quote:
                    tag = True
                elif c == '>' and not quote:
                    tag = False
                elif (c == '"' or c == "'") and tag:
                    quote = not quote
                elif not tag:
                    out = out + c
    return out


iExlFilename = 'PS00'
iExl=load_workbook(filename=iExlFilename+'.xlsx')
sh00 = iExl['00']

tree = parse('ruby_ps_00_shk_02.xml')
root = tree.getroot()
sce = root.findall("Element")
# subele = [x.find("SubElements") for x in sce]
nsce = 1
for x in sce:    
# for x in [sce[0]]:    
    #copy the template sheet to a new one
    #Fill each cell
    scename = x.get('Name')
    scegoal = x.get('Goal')
    sceid = x.get('SCEID')
    print(sceid,scename, scegoal)

    sh = iExl.copy_worksheet(sh00)
    sh.title = "PS-{:2d}".format(nsce)
    sh.cell(1,1).value = "PS-{:2.0f} {:s}".format(nsce,scename)
    #incase sceid is None
    # sh.title = "PS-{:2d}".format(sceid) 
    # sh.cell(1,1).value = "PS={:2.0f} {:s}".format(sceid,scename)
    sh.cell(3,1).value = scegoal    
    
    # print("<H1> {:s} </H1>".format(scename))

    Rows = np.array([6,10, 13, 16, 19])    
    ses = x.find('SubElements').getchildren()
    for se in ses:
        sename = se.get('Name')
        rse,rf,ra,rs,rd = Rows
        print(rse, sename)

        merged_cells_range = sh.merged_cells.ranges
        for merged_cell in merged_cells_range:
            if merged_cell.top[0][0] > rse:
                merged_cell.shift(0,1)
        sh.insert_rows(rse)
        sh.merge_cells(start_row=rse, start_column=1, end_row=rse, end_column=6)
        sh.cell(rse,1).value = sename
        Rows += 1

    print("<H2> {:s} </H2>".format("Functionality"))
    fs = x.find('Functionality').getchildren()
    for f in fs:                
        fref = f.get('Ref')        
        ffunction = f.get('Function')
        fcriteria = f.get('Criteria')
        fassurance = f.get('Assurance')
        print(fref, ffunction, fcriteria)
        rse,rf,ra,rs,rd = Rows        
        merged_cells_range = sh.merged_cells.ranges
        for merged_cell in merged_cells_range:
            if merged_cell.top[0][0] > rf:
                merged_cell.shift(0,1)
        sh.insert_rows(rf)
        sh.merge_cells(start_row=rf, start_column=2, end_row=rf, end_column=3)
        sh.cell(rf,1).value = fref
        sh.cell(rf,2).value = ffunction
        sh.cell(rf,4).value = fcriteria
        sh.cell(rf,5).value = fassurance
        Rows += 1
    
    # print("<H2> {:s} </H2>".format("Availability"))
    avs = x.find('Availability').getchildren()
    for f in avs:                
        ael = remove_html_markup(f.get('SubElement'))
        aa = f.get('AvailabilityAssurance')
        ac = f.get('AvailabilityCriterion')
        aref = f.get('AvID')        
        averif = f.get('AvailabilityVerification')        
        print(aref, ael,aa, ac)
        rse,rf,ra,rs,rd = Rows        
        merged_cells_range = sh.merged_cells.ranges
        for merged_cell in merged_cells_range:
            if merged_cell.top[0][0] > ra:
                merged_cell.shift(0,1)
        sh.insert_rows(ra)
        sh.merge_cells(start_row=ra, start_column=2, end_row=ra, end_column=3)
        sh.cell(ra,1).value = aref
        sh.cell(ra,2).value = ael
        sh.cell(ra,4).value = ac
        sh.cell(ra,5).value = aa
        sh.cell(ra,6).value = averif
        Rows += 1
    
    # print("<H2> {:s} </H2>".format("Survivability"))
    svs = x.find('Survivability').getchildren()
    for f in svs:                
        sref = f.get('SurvID')        
        smah = f.get('MAH')
        sel = remove_html_markup(f.get('SubElement'))
        sc = f.get('SurvCriterion')
        sa = f.get('SurvAssurance')
        sverif = f.get('SurvVerification')
        print(sref, smah, sel, sa, sc)    
        rse,rf,ra,rs,rd = Rows        
        merged_cells_range = sh.merged_cells.ranges
        for merged_cell in merged_cells_range:
            if merged_cell.top[0][0] > rs:
                merged_cell.shift(0,1)
        sh.insert_rows(rs)
        
        sh.cell(rs,1).value = sref
        sh.cell(rs,2).value = smah
        sh.cell(rs,3).value = sel
        sh.cell(rs,4).value = sc
        sh.cell(rs,5).value = sa
        sh.cell(rs,6).value = sverif
        Rows += 1
    
    svs = x.find('Dependencies').getchildren() #DependentSCE
    for f in svs:                
        dsce = f.get('DepSCEName')        
        ddes = f.get('DepDescription')
        
        print(dsce, ddes)    
        rse,rf,ra,rs,rd = Rows        
        merged_cells_range = sh.merged_cells.ranges
        for merged_cell in merged_cells_range:
            if merged_cell.top[0][0] > rd:
                merged_cell.shift(0,1)
        sh.merge_cells(start_row=rd, start_column=2, end_row=rd, end_column=6)
        sh.insert_rows(rd)
        
        sh.cell(rd,1).value = remove_html_markup(dsce)
        sh.cell(rd,2).value = remove_html_markup(ddes)
        
        Rows += 1
    
    nsce += 1


reflists = root.findall("ReferenceList")
# subele = [x.find("SubElements") for x in sce]
shrr = iExl['RulesReg']
rrs =     reflists[0].getchildren()
nrr = 1
for x in rrs:
    shrr.cell(nrr+2,1).value = x.get('ID')
    shrr.cell(nrr+2,2).value = x.get('AUTHOR')
    shrr.cell(nrr+2,3).value = remove_html_markup(x.get('DOCNO'))
    shrr.cell(nrr+2,4).value = x.get('REV')
    shrr.cell(nrr+2,5).value = remove_html_markup(x.get('TITLE'))
    nrr += 1

shid = iExl['InputDocs']
ids =     reflists[1].getchildren()
nid = 1
for x in ids:
    shid.cell(nid+2,1).value = x.get('ID')
    shid.cell(nid+2,2).value = x.get('AUTHOR')
    shid.cell(nid+2,3).value = remove_html_markup(x.get('DOCNO'))
    shid.cell(nid+2,4).value = x.get('REV')
    shid.cell(nid+2,5).value = remove_html_markup(x.get('TITLE'))
    nid += 1


""" 
refs = root.findall("ReferenceList")
rr = refs[0].getchildren()
for r in rr:    
    rauthor = r.get('AUTHOR')
    rdocno = r.get('DOCNO')
    rdocrev = r.get('REV')
    rdocttl = r.get('TITLE')
    print(rauthor, rdocno, rdocrev, rdocttl)

dd = refs[1].getchildren()
for r in dd:    
    rauthor = r.get('AUTHOR')
    rdocno = r.get('DOCNO')
    rdocrev = r.get('REV')
    rdocttl = r.get('TITLE')
    print(rauthor, rdocno, rdocrev, rdocttl)
 """

iExl.save('PS_'+time.strftime("%Y%m%d-%H%M%S")+'.xlsx')